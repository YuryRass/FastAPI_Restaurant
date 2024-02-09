from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings
from app.utils.json_shemas import JsonDish, JsonMenu, JsonSubmenu


class ExcelReader:
    """Чтение данных из Excel файла."""

    def __init__(self) -> None:
        self.sheet: Worksheet = load_workbook(filename=settings.EXCEL_PATH).active
        self.menus: list = []

    def get_menus(self) -> list[JsonMenu]:
        """Возвращает заполненную json структуру меню ресторана."""
        self.__reader()
        return self.menus

    def __reader(self) -> None:
        """
        Считывает данные всех меню ресторана из excel файла
        и сохраняет их в json структуре.
        """
        for row in self.sheet.iter_rows(values_only=True):
            # Меню
            if row[0] is not None:
                self.menus.append(
                    JsonMenu(
                        id=row[0],
                        title=row[1],
                        description=row[2],
                        submenus=[],
                    ).model_dump()
                )
            # Подменю
            elif row[1] is not None:
                self.menus[-1]['submenus'].append(
                    JsonSubmenu(
                        id=row[1],
                        title=row[2],
                        description=row[3],
                        dishes=[],
                    ).model_dump()
                )
            # Блюда
            else:
                # получаем скидку на блюдо
                discount = None
                if len(row) == 7:
                    discount = self.__get_discount(row[6])

                self.menus[-1]['submenus'][-1]['dishes'].append(
                    JsonDish(
                        id=row[2],
                        title=row[3],
                        description=row[4],
                        price=row[5] * discount if discount else row[5],
                    ).model_dump()
                )

    def __get_discount(self, discount: str | None) -> int | None:
        """Проверка и получение скидки на блюдо."""
        if discount and discount.isdigit() and 0 <= int(discount) <= 100:
            return int(discount)
        return None
